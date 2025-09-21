# ==== app.py 最頂部 ====
import os, sys
os.environ["STREAMLIT_GLOBAL_DEVELOPMENTMODE"] = "false"
os.environ["STREAMLIT_BROWSER_GATHERUSAGESTATS"] = "false"


import ctypes ,threading , subprocess
import io , socket
import streamlit as st
import sqlite3, io, datetime as dt, os, base64
import pandas as pd
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


def get_desktop():
    if sys.platform == "win32":
        # 0x0000 = CSIDL_DESKTOPDIRECTORY
        buf = ctypes.create_unicode_buffer(260)
        ctypes.windll.shell32.SHGetFolderPathW(None, 0x0000, None, 0, buf)
        return buf.value
    else:
        return os.path.join(os.path.expanduser("~"), "Desktop")
    
def get_output_dir(room: str, period: str) -> str:
    y, m = period.split("-")
    desktop = get_desktop()
    outdir = os.path.join(desktop, f"{y}年", f"{int(m)}月", room)
    os.makedirs(outdir, exist_ok=True)
    return outdir

def save_bytes(path: str, data: bytes):
    with open(path, "wb") as f:
        f.write(data)

def resource_base():
    # PyInstaller onefile 展開到臨時資料夾 sys._MEIPASS
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


# ---- 可選：HTML→PDF/PNG ----
try:
    import jinja2, pdfkit
    from pdf2image import convert_from_bytes
    HAS_IMG_EXPORT = True
except Exception:
    HAS_IMG_EXPORT = False

DB = "rent.db"

# ============== 工具 ==============
def round_int_half_up(x):
    """四捨五入到整數（傳統四捨五入）"""
    return int(Decimal(x).quantize(0, rounding=ROUND_HALF_UP))

def parse_int(s: str, default=0):
    """把文字輸入轉成整數；允許 1,000；空值→default"""
    if s is None: return default
    s = str(s).strip().replace(",", "")
    if s == "": return default
    try:
        return int(Decimal(s))
    except:
        return default

def int_input(label, default=0, key=None, help=None, placeholder=None):
    """文字輸入的整數欄位（無上下箭頭）"""
    raw = st.text_input(label, value=str(default), key=key, help=help, placeholder=placeholder)
    val = parse_int(raw, default=default)
    if raw and (str(val) != str(raw).strip().replace(",", "")):
        st.caption(f"已自動將「{label}」轉為整數：{val}")
    return val

# 人民幣中文大寫（整數部分）
def rmb_upper_int(n: int) -> str:
    if n == 0:
        return "零元整"
    digits = "零壹貳叁肆伍陸柒捌玖"
    units1 = ["", "拾", "佰", "仟"]
    units2 = ["", "萬", "億", "兆"]
    s = ""
    i = 0
    while n > 0:
        part = n % 10000
        if part != 0:
            part_str = ""
            zero_flag = False
            for j in range(4):
                d = part % 10
                if d == 0:
                    if not zero_flag and part_str:
                        part_str = "零" + part_str
                    zero_flag = True
                else:
                    part_str = digits[d] + units1[j] + part_str
                    zero_flag = False
                part //=10
                if part == 0 and j < 3:
                    break
            part_str = part_str.rstrip("零")
            s = part_str + units2[i] + s
        else:
            if not s.startswith("零") and s != "":
                s = "零" + s
        n //= 10000
        i += 1
    s = s.replace("零零", "零").rstrip("零")
    return s + "元整"

def spaced(text: str) -> str:
    """在中文大寫之間加空格，易讀可對齊"""
    return " ".join(list(text))

# ============== DB ==============
def init_db():
    with sqlite3.connect(DB) as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS bills_final(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          room TEXT, period TEXT, issue_date TEXT,
          water_prev INTEGER, water_curr INTEGER,
          elec_prev  INTEGER, elec_curr  INTEGER,
          car_prev   INTEGER, car_curr   INTEGER,
          water_used INTEGER, elec_used INTEGER, car_used INTEGER,
          water_fee  INTEGER, elec_fee  INTEGER, car_fee  INTEGER,
          rent       INTEGER, trash_fee INTEGER, network_fee INTEGER,
          other_label TEXT, other_fee INTEGER,
          utilities_sub INTEGER, total INTEGER,
          rmb_upper TEXT,
          note TEXT, created_at TEXT
        )
        """)

def insert_row(payload):
    with sqlite3.connect(DB) as con:
        con.execute("""
        INSERT INTO bills_final
        (room, period, issue_date,
         water_prev, water_curr, elec_prev, elec_curr, car_prev, car_curr,
         water_used, elec_used, car_used,
         water_fee, elec_fee, car_fee,
         rent, trash_fee, network_fee, other_label, other_fee,
         utilities_sub, total, rmb_upper, note, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            payload["room"], payload["period"], payload["issue_date"],
            payload["water_prev"], payload["water_curr"],
            payload["elec_prev"],  payload["elec_curr"],
            payload["car_prev"],   payload["car_curr"],
            payload["water_used"], payload["elec_used"], payload["car_used"],
            payload["water_fee"],  payload["elec_fee"],  payload["car_fee"],
            payload["rent"], payload["trash_fee"], payload["network_fee"],
            payload["other_label"], payload["other_fee"],
            payload["utilities_sub"], payload["total"], payload["rmb_upper"],
            payload["note"], dt.datetime.now().isoformat(timespec="seconds")
        ))

def load_history():
    with sqlite3.connect(DB) as con:
        return pd.read_sql_query(
            "SELECT id, room, period, total, rmb_upper, issue_date, created_at, note FROM bills_final ORDER BY id DESC",
            con
        )

# ============== Excel（美化 + 段落 + 收款碼） ==============
def pil_to_temp_png(uploaded_file):
    """把上傳的檔案標準化成 png bytes，避免 Excel 插圖時因格式出錯"""
    if uploaded_file is None:
        return None
    img = PILImage.open(uploaded_file).convert("RGBA")
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    bio.seek(0)
    return bio

def make_excel(row: dict, wx_qr=None, ali_qr=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    import io, math

    wb = Workbook()
    ws = wb.active
    ws.title = "收據"

    # ===== 樣式 =====
    header_fill = PatternFill("solid", fgColor="DCE6F1")   # 表頭淡藍灰
    title_fill  = PatternFill("solid", fgColor="BDD7EE")   # 標題藍
    para_fill   = PatternFill("solid", fgColor="FCE4D6")   # 段落淡橘
    sub_fill    = PatternFill("solid", fgColor="E2EFDA")   # 小計綠
    total_fill  = PatternFill("solid", fgColor="FCE4D6")   # 總計橘
    badge_fill  = PatternFill("solid", fgColor="EEECE1")   # 小標
    bold = Font(bold=True)
    big_bold = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===== 頂部「租戶您好」大段落（放最上面）=====
    period_cn = row["period"].replace("-", "年") + "月"
    para = (
        f"{row['room']}号房租户您好！你{period_cn}的租金为{row['rent']:.2f}元；"
        f"{period_cn}水、电费及费用{row['utilities_sub']:.2f}元；垃圾费为 {row['trash_fee']} 元/月。"
        f"合计(四舍五入){row['total']:.2f}元，（大写：{' '.join(list(row['rmb_upper']))}）。"
        f"请于本月5号前缴纳。水电用量明细如下表格。你可以通过以下微信或支付宝的二维码对我进行付款。"
    )
    # 估算段落需要的行高（把行數更保守、行高更大）
    chars_per_line = 32        # 每行字數估值（越小越高）
    est_lines = max(4, math.ceil(len(para) / chars_per_line))
    ws.merge_cells("A1:H3")
    c = ws["A1"]; c.value = para
    c.fill = para_fill; c.font = big_bold; c.alignment = wrap_left; c.border = border
    line_height = 24           # 每行高度 pt（加大）
    ws.row_dimensions[1].height = est_lines * line_height
    ws.row_dimensions[2].height = 2
    ws.row_dimensions[3].height = 2

    # ===== 標題（段落下方一行）=====
    ws.merge_cells("A4:H4")
    c = ws["A4"]; c.value = "房租 / 水電等收費收據"
    c.font = big_bold; c.fill = title_fill; c.alignment = center; c.border = border

    # ===== 基本資訊 =====
    r = 5
    for k, v in [("房號", row["room"]), ("期間", row["period"]),
                 ("開票日期", row["issue_date"]), ("備註", row["note"] or "")]:
        ws[f"A{r}"] = k
        ws[f"A{r}"].font = bold
        ws[f"A{r}"].fill = header_fill
        ws[f"A{r}"].alignment = center
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws[f"B{r}"] = v
        ws[f"B{r}"].alignment = left
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # ===== 表頭 =====
    r += 1
    head_row = r
    ws.append(["項目","上月","本月","實際用量","金額","其他","金額"])
    for col in range(1, 8):
        cell = ws.cell(row=head_row, column=col)
        cell.fill = header_fill; cell.font = bold; cell.alignment = center; cell.border = border

    # ===== 表體 =====
    def append_row(vals, fill=None):
        """直接追加一列（無合併）"""
        nonlocal r
        r += 1
        ws.append(vals)
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            if fill: cell.fill = fill
            cell.border = border
            cell.alignment = center if col in (2,3,4,5,6,8) else left

    def append_simple_line(label, amount, fill=None):
        nonlocal r
        r += 1
        # A 標題
        ws.cell(row=r, column=1, value=label)
        # B~D 合併（上月/本月/用量不適用）
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=2).value = ""
        # E 金額
        ws.cell(row=r, column=5, value=amount)
        # F~G 合併（其他）
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        ws.cell(row=r, column=6).value = ""
        # 邊框/填色照舊


    def append_uppercase_line(label, text, fill=None):
        nonlocal r
        r += 1
        ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        # E~G 放中文大寫（避免遮擋）
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        ws.cell(row=r, column=5, value=text)


    append_row(["水費", row["water_prev"], row["water_curr"], row["water_used"], row["water_fee"], "", ""])
    append_row(["房電費", row["elec_prev"], row["elec_curr"], row["elec_used"], row["elec_fee"], "", ""])
    append_row(["車房電費", row["car_prev"], row["car_curr"], row["car_used"], row["car_fee"], "", ""])

    append_simple_line("租金", row["rent"])
    append_simple_line("垃圾費", row["trash_fee"])
    append_simple_line("網絡維護", row["network_fee"])
    if row["other_fee"]:
        append_simple_line(row["other_label"] or "其他", row["other_fee"])

    append_simple_line("水電小計", row["utilities_sub"], fill=sub_fill)
    append_simple_line("總金額（¥）", row["total"], fill=total_fill)
    append_uppercase_line("（大寫）", row["rmb_upper"], fill=total_fill)

    # 欄寬（把第 6 欄加寬，避免長字被擋）
    widths = [14,12,12,14,20,12,12]  # A..G
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


    # ===== 收款碼區塊 =====
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws.cell(row=r, column=1)
    c.value = "掃碼付款"; c.alignment = center; c.font = big_bold; c.fill = badge_fill; c.border = border

    img_row = r + 1
    ws.row_dimensions[img_row].height = 140
    if wx_qr:
        xl_img = XLImage(wx_qr); xl_img.width, xl_img.height = 180, 180
        ws.add_image(xl_img, f"B{img_row}")
    if ali_qr:
        xl_img2 = XLImage(ali_qr); xl_img2.width, xl_img2.height = 180, 180
        ws.add_image(xl_img2, f"F{img_row}")

    label_row = img_row + 10
    if wx_qr:
        ws.merge_cells(start_row=label_row, start_column=2, end_row=label_row, end_column=4)
        c = ws.cell(row=label_row, column=2)
        c.value = "微信收款碼"; c.alignment = center; c.font = bold; c.fill = header_fill; c.border = border
    if ali_qr:
        ws.merge_cells(start_row=label_row, start_column=6, end_row=label_row, end_column=8)
        c = ws.cell(row=label_row, column=6)
        c.value = "支付寶收款碼"; c.alignment = center; c.font = bold; c.fill = header_fill; c.border = border

    # ===== 全表粗體（安全版）=====
    from openpyxl.styles import Font
    for rr in range(1, ws.max_row + 1):
        for cc in range(1, 9):
            cell = ws.cell(row=rr, column=cc)
            curr = cell.font
            curr_size = getattr(curr, "size", getattr(curr, "sz", 11))
            if curr_size and float(curr_size) >= 14:
                continue
            new_font = Font(
                name=getattr(curr, "name", "Calibri"),
                size=curr_size,
                bold=True,
                italic=getattr(curr, "italic", False),
                vertAlign=getattr(curr, "vertAlign", None),
                underline=getattr(curr, "underline", "none"),
                strike=getattr(curr, "strike", False),
                color=getattr(curr, "color", None),
                charset=getattr(curr, "charset", None),
                scheme=getattr(curr, "scheme", None),
            )
            cell.font = new_font
    1                                                                                                                                                               
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return f'{row["period"]}_{row["room"]}.xlsx', bio

# ============== HTML→PNG（可選） ==============
HTML_TEMPLATE = """
<!doctype html>
<html lang="zh-Hans">
<meta charset="utf-8"/>
<style>
body { font-family: -apple-system, blinkmacsystemfont, "Segoe UI", Roboto, "Noto Sans CJK SC","PingFang SC","Microsoft YaHei", Arial, sans-serif; margin: 24px; }
.box { max-width: 900px; margin: 0 auto; border: 1px solid #ddd; padding: 18px; }
h1 { text-align:center; font-size: 20px; background:#BDD7EE; padding:10px 0; margin:0 0 12px 0;}
p.lead { background:#f7f7f7; padding:10px; line-height:1.7; }
table { width:100%; border-collapse: collapse; margin-top:10px; font-size:14px;}
th, td { border:1px solid #bbb; padding:6px 8px; text-align:center;}
th { background:#DCE6F1; }
tfoot td { background:#FCE4D6; font-weight:bold; }
.badge { display:inline-block; padding:2px 8px; background:#E2EFDA; border-radius:8px; font-size:12px; }
.qrs { display:flex; gap:20px; justify-content:space-around; margin-top:14px; }
.qrs .col { text-align:center; }
.qrs img { width:180px; height:180px; object-fit:contain; border:1px solid #ddd; }
.meta { margin-top:8px; }
</style>
<div class="box">
  <h1>房租 / 水電等收費收據</h1>
  <p class="lead">
    {{ room }}号房租户您好！你{{ period_cn }}的租金为{{ rent_fmt }}元；{{ period_cn }}水、电费及费用{{ utilities_fmt }}元；
    垃圾费为 {{ trash }} 元/月。合计(四舍五入){{ total_fmt }}元，（大写：{{ rmb_upper_spaced }}）。请于本月5号前缴纳。
    水电用量明细如下表格。你可以通过以下微信或支付宝的二维码对我进行付款。
  </p>

  <div class="meta">
    <span class="badge">房號：{{ room }}</span>
    <span class="badge">期間：{{ period }}</span>
    <span class="badge">開票日期：{{ issue_date }}</span>
    {% if note %}<span class="badge">備註：{{ note }}</span>{% endif %}
  </div>

  <table>
    <thead>
      <tr><th>項目</th><th>上月</th><th>本月</th><th>實際用量</th><th>金額</th></tr>
    </thead>
    <tbody>
        <tr><td>水費</td><td>{{ water_prev }}</td><td>{{ water_curr }}</td><td>{{ water_used }}</td><td>{{ water_fee }}</td></tr>
        <tr><td>房電費</td><td>{{ elec_prev }}</td><td>{{ elec_curr }}</td><td>{{ elec_used }}</td><td>{{ elec_fee }}</td></tr>
        <tr><td>車房電費</td><td>{{ car_prev }}</td><td>{{ car_curr }}</td><td>{{ car_used }}</td><td>{{ car_fee }}</td></tr>
        <tr><td>租金</td><td colspan="3"></td><td>{{ rent }}</td></tr>
        <tr><td>垃圾費</td><td colspan="3"></td><td>{{ trash }}</td></tr>
        <tr><td>網絡維護</td><td colspan="3"></td><td>{{ network }}</td></tr>
        {% if other_fee and other_fee != 0 %}
        <tr><td>{{ other_label or "其他" }}</td><td colspan="3"></td><td>{{ other_fee }}</td></tr>
        {% endif %}
    </tbody>
    <tfoot>
        <tr><td>水電小計</td><td colspan="3"></td><td>{{ utilities_sub }}</td></tr>
        <tr><td>總金額（¥）</td><td colspan="3"></td><td>{{ total }}</td></tr>
        <tr><td>（大寫）</td><td colspan="3"></td><td>{{ rmb_upper }}</td></tr>
    </tfoot>
  </table>

  <div class="qrs">
    {% if wx_qr %}<div class="col"><div>微信收款碼</div><img src="data:image/png;base64,{{ wx_qr }}"/></div>{% endif %}
    {% if ali_qr %}<div class="col"><div>支付寶收款碼</div><img src="data:image/png;base64,{{ ali_qr }}"/></div>{% endif %}
  </div>
</div>
"""

def render_png_via_html(row, wx_qr_bytes=None, ali_qr_bytes=None):
    if not HAS_IMG_EXPORT:
        return None, None

    # 取得目前 app.py 所在目錄
    base_dir = resource_base()
    wkhtml_path = os.path.join(base_dir, "wkhtmltopdf", "bin", "wkhtmltopdf.exe")
    poppler_bin = os.path.join(base_dir, "poppler-25.07.0", "Library", "bin")

    # 檢查檔案是否存在
    if not os.path.isfile(wkhtml_path):
        print("找不到 wkhtmltopdf.exe:", wkhtml_path)
        return None, None
    if not os.path.isfile(os.path.join(poppler_bin, "pdftoppm.exe")):
        print("找不到 pdftoppm.exe:", poppler_bin)
        return None, None

    # 1) 渲染 HTML
    env = jinja2.Environment(autoescape=True)
    tpl = env.from_string(HTML_TEMPLATE)
    def b64(b):
        import base64
        return base64.b64encode(b).decode("ascii") if b else None

    html = tpl.render(
        room=row["room"],
        period=row["period"],
        period_cn=row["period"].replace("-", "年") + "月",
        issue_date=row["issue_date"],
        note=row["note"],
        rent=row["rent"], trash=row["trash_fee"], network=row["network_fee"],
        rent_fmt=f"{row['rent']:.2f}", utilities_fmt=f"{row['utilities_sub']:.2f}", total_fmt=f"{row['total']:.2f}",
        water_prev=row["water_prev"], water_curr=row["water_curr"], water_used=row["water_used"], water_fee=row["water_fee"],
        elec_prev=row["elec_prev"], elec_curr=row["elec_curr"], elec_used=row["elec_used"], elec_fee=row["elec_fee"],
        car_prev=row["car_prev"], car_curr=row["car_curr"], car_used=row["car_used"], car_fee=row["car_fee"],
        other_label=row["other_label"], other_fee=row["other_fee"],
        utilities_sub=row["utilities_sub"], total=row["total"],
        rmb_upper=row["rmb_upper"], rmb_upper_spaced=" ".join(list(row["rmb_upper"])),
        wx_qr=b64(wx_qr_bytes), ali_qr=b64(ali_qr_bytes),
    )

    # 2) HTML -> PDF
    config = pdfkit.configuration(wkhtmltopdf=wkhtml_path)
    options = {
        "quiet": "",
        "disable-smart-shrinking": "",
        "print-media-type": ""
    }
    try:
        pdf_bytes = pdfkit.from_string(
            html,
            False,
            configuration=config,
            options=options   # ✅ 這裡加上 options
        )
    except Exception as e:
        print("wkhtmltopdf 轉 PDF 失敗:", e)
        return None, None


    # 3) PDF -> PNG
    try:
        images = convert_from_bytes(pdf_bytes, dpi=200, poppler_path=poppler_bin)
        buf = io.BytesIO()
        images[0].save(buf, format="PNG")
        buf.seek(0)
        return "receipt.png", buf
    except Exception as e:
        print("poppler 轉 PNG 失敗:", e)
        return None, None

# ============== App UI ==============
st.set_page_config(page_title="出租屋結算（最終版）", page_icon="🧾", layout="centered")
init_db()

st.title("🧾 出租屋結算")

with st.form("bill_form", clear_on_submit=False):
    # 基本資料
    c1, c2 = st.columns(2)
    room   = c1.text_input("房號", "")
    period = c2.text_input("結算月份（yyyy-mm）", dt.date.today().strftime("%Y-%m"))
    issue_date = st.date_input("開票日期", dt.date.today())
    st.divider()

    # 讀數：每項兩欄並列
    st.subheader("讀數(上月/本月)")
    cc1, cc2 = st.columns(2)
    with cc1:
        st.markdown("**水費**")
        water_prev = int_input("上月水表", 0, key="w_prev", placeholder="0")
    with cc2:
        st.markdown("**.**")
        water_curr = int_input("本月水表", 0, key="w_curr", placeholder="0")

    cc1, cc2 = st.columns(2)
    with cc1:
        st.markdown("**房電費**")
        elec_prev  = int_input("上月房電", 0, key="e_prev", placeholder="0")
    with cc2:
        st.markdown("**.**")
        elec_curr  = int_input("本月房電", 0, key="e_curr", placeholder="0")

    cc1, cc2 = st.columns(2)
    with cc1:
        st.markdown("**車房電費**")
        car_prev   = int_input("上月車房電", 0, key="c_prev", placeholder="0")
    with cc2:
        st.markdown("**.**")
        car_curr   = int_input("本月車房電", 0, key="c_curr", placeholder="0")

    st.divider()

    # 其他費用
    st.subheader("其他費用")
    cc1, cc2 = st.columns(2)
    rent = int_input("租金(¥)", 0, key="rent", placeholder="0")
    trash_fee = int_input("垃圾費(¥)", 10, key="trash", placeholder="10")

    cc1, cc2 = st.columns(2)
    network_fee = int_input("網絡維護(¥)", 30, key="net", placeholder="30")
    other_label = st.text_input("其他項目名稱（可空）", "", placeholder="例如：維修費")
    other_fee = int_input("其他項目金額(¥)", 0, key="other", placeholder="0")

    note = st.text_input("備註", "", placeholder="可空")

    st.divider()
    st.subheader("收款碼（記得上傳 PNG/JPG）")
    wx_file  = st.file_uploader("微信收款碼", type=["png","jpg","jpeg"], key="wx")
    ali_file = st.file_uploader("支付寶收款碼", type=["png","jpg","jpeg"], key="ali")

    submitted = st.form_submit_button("計算、保存並生成")

if submitted:
    # 用量（不可負）
    water_used = max(0, int(water_curr) - int(water_prev))
    elec_used  = max(0, int(elec_curr)  - int(elec_prev))
    car_used   = max(0, int(car_curr)   - int(car_prev))

    # 金額（四捨五入為整數）
    water_fee = round_int_half_up(water_used * 5)
    elec_fee  = round_int_half_up(elec_used  * 1.2)
    car_fee   = round_int_half_up(car_used   * 1.2)

    utilities_sub = int(water_fee + elec_fee + car_fee)
    total = int(utilities_sub + rent + trash_fee + network_fee + other_fee)
    rmb_upper = rmb_upper_int(total)

    row = {
        "room": room, "period": period, "issue_date": issue_date.isoformat(),
        "water_prev": water_prev, "water_curr": water_curr,
        "elec_prev": elec_prev, "elec_curr": elec_curr,
        "car_prev": car_prev, "car_curr": car_curr,
        "water_used": water_used, "elec_used": elec_used, "car_used": car_used,
        "water_fee": water_fee, "elec_fee": elec_fee, "car_fee": car_fee,
        "rent": rent, "trash_fee": trash_fee, "network_fee": network_fee,
        "other_label": other_label, "other_fee": other_fee,
        "utilities_sub": utilities_sub, "total": total, "rmb_upper": rmb_upper,
        "note": note
    }

    insert_row(row)
    # ========= 改這裡（取 bytes，分別用新 BytesIO 供 Excel 與 PNG 使用） =========
    # 先把上傳圖轉成 PNG，再取出 bytes
    wx_png = pil_to_temp_png(wx_file)
    ali_png = pil_to_temp_png(ali_file)
    wx_bytes = wx_png.getvalue() if wx_png else None
    ali_bytes = ali_png.getvalue() if ali_png else None

    # 傳給 Excel：用新的 BytesIO 副本，避免 openpyxl 讀完後把流關掉
    fname_xlsx, xlsx_bytes = make_excel(
        row,
        wx_qr=io.BytesIO(wx_bytes) if wx_bytes else None,
        ali_qr=io.BytesIO(ali_bytes) if ali_bytes else None
    )

    st.download_button(
        "下載 Excel 收據",
        data=xlsx_bytes,
        file_name=fname_xlsx,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # PNG 收據（可選功能）：直接用 bytes（不是已被關閉的流）
    if HAS_IMG_EXPORT:
        fname_png, png_buf = render_png_via_html(
            row,
            wx_qr_bytes=wx_bytes,
            ali_qr_bytes=ali_bytes
        )
        if fname_png and png_buf:
            st.download_button("下載 PNG 收據（圖片）", data=png_buf, file_name=fname_png, mime="image/png")
        else:
            st.info("如要輸出 PNG，請安裝 wkhtmltopdf + poppler，或檢查 PATH。")
    else:
        st.info("圖片輸出未啟用。若需要 PNG：pip install jinja2 pdfkit pdf2image 並安裝 wkhtmltopdf、poppler。")
    # ========= 改到這裡結束 =========
    # ---- 自動保存到桌面/年份/月/房號 ----
    outdir = get_output_dir(room=row["room"], period=row["period"])

    # Excel 檔落地
    xlsx_path = os.path.join(outdir, f'{row["period"]}_{row["room"]}.xlsx')
    save_bytes(xlsx_path, xlsx_bytes.getvalue())

    # PNG（若有啟用並成功）
    if HAS_IMG_EXPORT:
        fname_png, png_buf = render_png_via_html(row, wx_qr_bytes=wx_bytes, ali_qr_bytes=ali_bytes)
        if fname_png and png_buf:
            png_path = os.path.join(outdir, f'{row["period"]}_{row["room"]}.png')
            save_bytes(png_path, png_buf.getvalue())
            st.success(f"圖片已保存到：{png_path}")
        else:
            st.info("未能生成 PNG（請確認本目錄有 wkhtmltopdf\\bin 與 poppler-*/Library/bin）。")
    else:
        st.info("圖片輸出未啟用。若需要 PNG：pip install jinja2 pdfkit pdf2image。")
        
    st.success(f"Excel 已保存到：{xlsx_path}")

st.subheader("📜 歷史記錄")
df = load_history()
st.dataframe(df)
